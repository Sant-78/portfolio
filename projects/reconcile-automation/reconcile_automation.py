
from __future__ import annotations
import argparse
import math

import re
import smtplib
from dataclasses import dataclass
from datetime import datetime, timedelta, date
from email.message import EmailMessage
from pathlib import Path
from typing import Iterable

from openpyxl import Workbook,load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, NamedStyle

import pandas as pd
import xlrd

TB_ENCODING = "latin1"
TB_NAME_PATTERN = "TrialBalRpt_*.csv"
ORACLE_NAME_PATTERN = "Flexcube_GL_Dump_*.xls"
DATE_FMT_TB = "%d%m%Y"
DATE_FMT_ORACLE = "%d-%m-%Y"
DATE_FMT_HEADER = "%Y-%m-%d"
DATE_FMT_OUTPUT = "%d-%b-%Y"

SMTP_SERVER = "smtp.office365.com"
SMTP_PORT = 587
SMTP_USER = "Santosh.Pal@caprihomeloans.com"
SMTP_PASSWORD = "xxxxx"


@dataclass(frozen=True)
class DayData:
    recon_date: datetime
    tb_asof_date: datetime
    tb_closing: pd.Series
    oracle_net: pd.Series


def parse_args() -> argparse.Namespace:
    today = datetime.today()
    parser = argparse.ArgumentParser(
        description="Automate Flexcube vs Oracle reconciliation into the provided template."
    )
    parser.add_argument(
        "--root",
        default=r"Z:\Accounts\Santosh.Pal\Flexcube_Report",
        help="Base working folder that contains year/month folders.",
    )
    parser.add_argument(
        "--year",
        type=int,
        default=today.year,
        help=f"Year folder, for example 2026. Default: {today.year}",
    )
    parser.add_argument(
        "--month",
        default=f"{today.month:02d}",
        help=f"Month folder, for example 05. Default: {today.month:02d}",
    )
    parser.add_argument(
        "--company",
        choices=["CGCL", "CGHFL", "ALL"],
        default="ALL",
        help="Run for one company or both.",
    )
    parser.add_argument(
        "--template",
        default=None,
        help="Optional template path. Defaults to <root>/<year>/<month>/Template.xlsx",
    )
    parser.add_argument(
        "--output-dir",
        default=None,
        help="Optional output folder. Defaults to <root>/<year>/<month>/Reconciliation_Output",
    )
    parser.add_argument(
        "--no-email",
        action="store_true",
        help="Generate the reports without sending the SMTP email.",
    )
    return parser.parse_args()


def month_root(root: Path, year: int, month: str) -> Path:
    return root / str(year) / f"{int(month):02d}"


def parse_tb_date(path: Path) -> datetime:
    return datetime.strptime(path.stem.split("_")[-1], DATE_FMT_TB)


def parse_oracle_date(path: Path) -> datetime:
    return datetime.strptime(path.stem.split("_")[-1], DATE_FMT_ORACLE)


def format_day_label(day: datetime) -> str:
    suffix = "th"
    if day.day % 10 == 1 and day.day % 100 != 11:
        suffix = "st"
    elif day.day % 10 == 2 and day.day % 100 != 12:
        suffix = "nd"
    elif day.day % 10 == 3 and day.day % 100 != 13:
        suffix = "rd"
    return f"{day.day}{suffix} {day.strftime('%b')} closing balance"


def read_tb_file(path: Path) -> pd.Series:
    df = pd.read_csv(path, encoding=TB_ENCODING, on_bad_lines="skip")
    if "GL Code" not in df.columns or "Closing Balance" not in df.columns:
        return pd.Series(dtype="float64")
    df["GL Code"] = pd.to_numeric(df["GL Code"], errors="coerce")
    df["Closing Balance"] = pd.to_numeric(df["Closing Balance"], errors="coerce").fillna(0.0)
    grouped = df.dropna(subset=["GL Code"]).groupby("GL Code", as_index=True)["Closing Balance"].sum()
    grouped.index = grouped.index.astype("Int64")
    return grouped.sort_index()


def read_oracle_file(path: Path) -> dict[date, pd.Series]: # type: ignore
    book = xlrd.open_workbook(str(path), on_demand=True)
    sheet_names = [
        name for name in book.sheet_names() if re.fullmatch(r"Sheet\d*", name)
    ]
    sheet_names = sorted(
        sheet_names,
        key=lambda name: int(m.group()) if (m := re.search(r"\d+", name)) is not None else 1,
    )

    def find_header_row(sheet: object) -> int:
        for row_idx in range(min(sheet.nrows, 30)):
            text = "|".join(str(v).upper() for v in sheet.row_values(row_idx)[:80])
            if "EFFECTIVE DATE" in text and "ACCOUNTED_DR" in text:
                return row_idx
        return -1

    all_rows: list[dict[str, object]] = []
    for sheet_name in sheet_names:
        sheet = book.sheet_by_name(sheet_name)
        header_row = find_header_row(sheet)
        start_row = header_row + 1 if header_row >= 0 else 0
        for row_idx in range(start_row, sheet.nrows):
            row = sheet.row_values(row_idx)
            if len(row) < 74:
                continue
            effective_date = str(row[2]).strip()
            legacy_code = row[73]
            accounted_dr = row[11] if len(row) > 11 else 0.0
            accounted_cr = row[12] if len(row) > 12 else 0.0
            if not effective_date or effective_date.upper() == "EFFECTIVE DATE":
                continue
            all_rows.append(
                {
                    "effective_date": effective_date,
                    "legacy_code": legacy_code,
                    "accounted_dr": accounted_dr,
                    "accounted_cr": accounted_cr,
                }
            )

    if not all_rows:
        return {} # type: ignore

    df = pd.DataFrame(all_rows)
    df["effective_date"] = pd.to_datetime(df["effective_date"], format="%d-%m-%Y", errors="coerce")
    df["legacy_code"] = pd.to_numeric(df["legacy_code"], errors="coerce")
    df["accounted_dr"] = pd.to_numeric(df["accounted_dr"], errors="coerce").fillna(0.0)
    df["accounted_cr"] = pd.to_numeric(df["accounted_cr"], errors="coerce").fillna(0.0)
    df = df.dropna(subset=["effective_date", "legacy_code"])
    df["oracle_net"] = df["accounted_cr"] - df["accounted_dr"]
    output: dict[date, pd.Series] = {} # type: ignore
    for effective_date, group in df.groupby(df["effective_date"].dt.date):
        grouped = group.groupby("legacy_code", as_index=True)["oracle_net"].sum()
        grouped.index = grouped.index.astype("Int64")
        output[effective_date] = grouped.sort_index()
    return output


def build_day_data(tb_dir: Path, oracle_dir: Path) -> list[DayData]:
    if not tb_dir.exists():
        raise FileNotFoundError(f"TB folder not found: {tb_dir}")
    if not oracle_dir.exists():
        raise FileNotFoundError(f"Oracle folder not found: {oracle_dir}")

    tb_files = sorted(tb_dir.glob(TB_NAME_PATTERN), key=parse_tb_date)
    oracle_files = sorted(oracle_dir.glob(ORACLE_NAME_PATTERN), key=parse_oracle_date)

    tb_by_date = {parse_tb_date(path).date(): path for path in tb_files}
    if not oracle_files:
        raise FileNotFoundError(f"No Oracle files found in {oracle_dir}")

    latest_oracle_file = oracle_files[-1]
    oracle_daily = read_oracle_file(latest_oracle_file) # type: ignore
    if not oracle_daily:
        raise FileNotFoundError(f"No Oracle transactions found in latest file: {latest_oracle_file.name}")

    cumulative_oracle: dict[date, pd.Series] = {} # type: ignore
    running = pd.Series(dtype="float64")
    for recon_day in sorted(oracle_daily): # type: ignore
        running = running.add(oracle_daily[recon_day], fill_value=0.0)
        cumulative_oracle[recon_day] = running.sort_index()

    day_data: list[DayData] = []
    for recon_day in sorted(cumulative_oracle):
        tb_asof_day = recon_day + timedelta(days=1)
        tb_path = tb_by_date.get(tb_asof_day)
        if tb_path is None:
            continue

        day_data.append(
            DayData(
                recon_date=datetime.combine(recon_day, datetime.min.time()),
                tb_asof_date=datetime.combine(tb_asof_day, datetime.min.time()),
                tb_closing=read_tb_file(tb_path),
                oracle_net=cumulative_oracle[recon_day],
            )
        )

    if not day_data:
        raise FileNotFoundError(
            f"No shifted TB/Oracle date pairs found in {tb_dir} and {oracle_dir}. "
            "Expected Oracle D to match TB D+1."
        )

    return day_data


def find_dates_with_diff(output_path: Path) -> tuple[list[str], list[str]]:
    wb = load_workbook(output_path, data_only=True)
    ws = wb[wb.sheetnames[0]]
    
    all_dates: list[str] = []
    diff_dates: list[str] = []
    
    for col_idx in range(3, ws.max_column + 1, 4):
        date_cell = ws.cell(1, col_idx)
        date_value = date_cell.value
        if date_value:
            if isinstance(date_value, datetime):
                date_str: str = date_value.strftime(DATE_FMT_OUTPUT)
            else:
                date_str = str(date_value)
            all_dates.append(date_str)
            diff_col = col_idx + 3
            max_row = ws.max_row
            has_diff = False
            for row_idx in range(3, max_row + 1):
                diff_value = ws.cell(row_idx, diff_col).value
                if diff_value is not None:
                    if isinstance(diff_value, (int, float)) and abs(float(diff_value)) >= 1.0:
                        has_diff = True
                        break
            if has_diff:
                diff_dates.append(date_str)
    
    return all_dates, diff_dates


def set_numeric_or_dash(cell, value: float | None) -> None:
    if value is None or (isinstance(value, float) and math.isnan(value)):
        cell.value = " -   "
        return
    numeric = round(float(value), 2)
    if abs(numeric) < 0.005:
        numeric = 0.0
    cell.value = numeric


def populate_template(
    output_path: Path,
    baseline_date: datetime,
    baseline_balances: pd.Series,
    days: list[DayData],
    existing_max_col: int = 2,
) -> None:
    if output_path.exists():
        wb = load_workbook(output_path)
    else:
        wb = Workbook()
        ws = wb.active
        # Set up headers for new file
        light_blue_fill = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")
        header_font = Font(bold=True)
        ws.cell(1, 1).value = None
        ws.cell(2, 1).value = "GL Code"
        ws.cell(1, 2).value = baseline_date
        ws.cell(2, 2).value = "Closing Balance"
        for row in [1, 2]:
            for col in [1, 2]:
                cell = ws.cell(row, col)
                cell.fill = light_blue_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center", vertical="center")
        data_style = NamedStyle(name="data_style")
        data_style.font = Font(bold=False)
        data_style.number_format = '#,##0.00;-#,##0.00'
        data_style.alignment = Alignment(horizontal="right", vertical="center")
        wb.add_named_style(data_style)
    
    ws = wb.active
    
    # Define styles
    light_blue_fill = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")
    header_font = Font(bold=True)
    data_style = NamedStyle(name="data_style")
    data_style.font = Font(bold=False)
    data_style.number_format = '#,##0.00;-#,##0.00'
    data_style.alignment = Alignment(horizontal="right", vertical="center")
    if "data_style" not in wb.named_styles:
        wb.add_named_style(data_style)

    base_balances = baseline_balances
    all_codes: list[int] = sorted(
        {
            int(code)
            for day in days
            for code in day.tb_closing.index.tolist() + day.oracle_net.index.tolist()
            if int(code) not in {66666666, 77777777}
        }
    )
    
    # Unmerge all merged cells in the worksheet before writing
    for merged_range in list(ws.merged_cells.ranges):
        ws.unmerge_cells(str(merged_range))

    # Clear stale data rows from a previous run
    if ws.max_row > 2:
        ws.delete_rows(3, ws.max_row - 2)
    
    # Set up headers
    ws.cell(1, 1).value = None
    ws.cell(2, 1).value = "GL Code"
    ws.cell(1, 2).value = baseline_date
    ws.cell(2, 2).value = "Closing Balance"
    
    # Apply header styling to first two columns
    for row in [1, 2]:
        for col in [1, 2]:
            cell = ws.cell(row, col)
            cell.fill = light_blue_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")

    # Set up date headers with merging and styling
    for idx, day in enumerate(days):
        start_col = 3 + (idx * 4)
        # Merge cells for date header (spanning 4 columns)
        ws.merge_cells(start_row=1, start_column=start_col, end_row=1, end_column=start_col + 3)
        
        # Set date value in merged cell
        date_cell = ws.cell(1, start_col)
        date_cell.value = day.recon_date
        date_cell.fill = light_blue_fill
        date_cell.font = header_font
        date_cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # Set sub-headers
        ws.cell(2, start_col).value = f"{format_day_label(day.tb_asof_date).replace('closing balance', 'Flexcube Closing balance')}"
        ws.cell(2, start_col + 1).value = "Flexcube_Net"
        ws.cell(2, start_col + 2).value = "Oracle_Net"
        ws.cell(2, start_col + 3).value = "Diff"
        
        # Apply header styling to sub-headers
        for offset in range(4):
            sub_header_cell = ws.cell(2, start_col + offset)
            sub_header_cell.fill = light_blue_fill
            sub_header_cell.font = header_font
            sub_header_cell.alignment = Alignment(horizontal="center", vertical="center")

    start_row = 3
    for offset, gl_code in enumerate(all_codes):
        row = start_row + offset
        base_value = float(base_balances.get(gl_code, 0.0))
        
        # GL Code column
        gl_code_cell = ws.cell(row, 1)
        gl_code_cell.value = gl_code
        gl_code_cell.fill = light_blue_fill
        gl_code_cell.font = header_font
        
        # Baseline balance column
        baseline_cell = ws.cell(row, 2)
        baseline_cell.value = base_value
        baseline_cell.style = "data_style"

        for idx, day in enumerate(days):
            start_col = 3 + (idx * 4)
            current_closing = day.tb_closing.get(gl_code)
            oracle_value = day.oracle_net.get(gl_code)

            # Flexcube Closing Balance
            closing_cell = ws.cell(row, start_col)
            if pd.isna(current_closing):
                set_numeric_or_dash(closing_cell, None)
            else:
                current_value = float(current_closing)
                set_numeric_or_dash(closing_cell, current_value)
            closing_cell.style = "data_style"

            # Flexcube_Net (change from baseline)
            flex_net_cell = ws.cell(row, start_col + 1)
            if pd.isna(current_closing):
                set_numeric_or_dash(flex_net_cell, None)
            else:
                current_value = float(current_closing)
                set_numeric_or_dash(flex_net_cell, current_value - base_value)
            flex_net_cell.style = "data_style"

            # Oracle_Net
            oracle_cell = ws.cell(row, start_col + 2)
            if pd.isna(oracle_value):
                set_numeric_or_dash(oracle_cell, None)
            else:
                set_numeric_or_dash(oracle_cell, float(oracle_value))
            oracle_cell.style = "data_style"

            # Diff (Flexcube_Net - Oracle_Net)
            diff_cell = ws.cell(row, start_col + 3)
            flex_value = ws.cell(row, start_col + 1).value
            oracle_cell_value = ws.cell(row, start_col + 2).value
            flex_num = float(flex_value) if isinstance(flex_value, (int, float)) else None
            oracle_num = float(oracle_cell_value) if isinstance(oracle_cell_value, (int, float)) else None
            if flex_num is not None or oracle_num is not None:
                diff_value = (flex_num or 0.0) - (oracle_num or 0.0)
                set_numeric_or_dash(diff_cell, diff_value)
            else:
                set_numeric_or_dash(diff_cell, None)
            diff_cell.style = "data_style"

    wb.save(output_path)


def run_company(root: Path, year: int, month: str, output_dir: Path, company: str) -> Path:
    base = month_root(root, year, month)
    tb_dir = base / "TB_Report" / company
    oracle_dir = base / "Oracle_Fustion_GL_Dump" / company
    tb_files = sorted(tb_dir.glob(TB_NAME_PATTERN), key=parse_tb_date)
    if not tb_files:
        raise FileNotFoundError(f"No TB files found in {tb_dir}")

    baseline_path = tb_files[0]
    baseline_date = parse_tb_date(baseline_path) - timedelta(days=1)
    baseline_balances = read_tb_file(baseline_path)
    days = build_day_data(tb_dir, oracle_dir)

    output_dir.mkdir(parents=True, exist_ok=True)
    output_path = output_dir / f"{company}_Reconciliation_{year}_{int(month):02d}.xlsx"
    populate_template(
        output_path=output_path,
        baseline_date=baseline_date,
        baseline_balances=baseline_balances,
        days=days,
    )
    return output_path


def send_reconciliation_email(
    output_files_by_company: dict[str, Path],
    report_date: str,
    all_dates_by_company: dict[str, list[str]],
    diff_dates_by_company: dict[str, list[str]],
) -> None:
    signature = "\n\nRegards,\nSantosh Pal\nExecutive - Finance & Accounts\nCapri Global Capital Limited\nGOREGAON 3\n\nNote: This is an auto-generated system email."

    for company, output_file in output_files_by_company.items():
        diff_dates = diff_dates_by_company.get(company, [])
        all_dates = all_dates_by_company.get(company, [])

        if diff_dates:
            subject = f"WARNING - {company} Flexcube TrialBalRpt v/s  Oracle GL Report Notification - {report_date}"
            unique_dates = sorted(set(diff_dates))
            body_lines = [
                "Dear Team,\n\n",
                "Please find attached the report for the Flexcube TrialBalRpt v/s Oracle GL Reconciliation today.\n\n",
                f"Differences found for {company} on the following dates:\n",
                f"{company} - {', '.join(unique_dates)}\n",
                signature,
            ]
        else:
            subject = f"SUCCESS - {company} Flexcube TrialBalRpt v/s  Oracle GL Report Notification - {report_date}"
            if all_dates:
                min_date = min(all_dates)
                max_date = max(all_dates)
                date_range = f"{min_date} to {max_date}"
            else:
                date_range = report_date
            body_lines = [
                "Dear Team,\n\n",
                "Please find attached the report for the Flexcube TrialBalRpt v/s Oracle GL Reconciliation today.\n\n",
                f"SUCCESS: All financial records match successfully between Flexcube TrialBalRpt and Oracle GL Report GL for {company} for date:\n",
                f"{company} - {date_range}\n",
                signature,
            ]

        msg = EmailMessage()
        msg["Subject"] = subject
        msg["From"] = SMTP_USER
        msg["To"] = "Santosh.Pal@caprihomeloans.com"     # corporateaccounts@capriglobal.in
        msg["CC"] = ""
        msg.set_content("".join(body_lines))

        with open(output_file, "rb") as f:
            msg.add_attachment(f.read(), maintype="application", subtype="vnd.openxmlformats-officedocument.spreadsheetml.sheet", filename=output_file.name)

        try:
            with smtplib.SMTP(SMTP_SERVER, SMTP_PORT) as server:
                server.starttls()
                server.login(SMTP_USER, SMTP_PASSWORD)
                server.send_message(msg)
            print(f"Email sent successfully via SMTP for {company}.")
        except Exception as smtp_error:
            print(f"SMTP error for {company}: {smtp_error}")
            print(f"Trying Outlook COM object fallback for {company}...")
            try:
                import win32com.client as win32
                outlook = win32.Dispatch("Outlook.Application")
                mail = outlook.CreateItem(0)
                mail.Subject = subject
                mail.To = "corporateaccounts@capriglobal.in"
                mail.CC = ""
                mail.Body = "".join(body_lines)
                mail.Attachments.Add(str(output_file.resolve()))
                mail.Send()
                print(f"Email sent successfully via Outlook for {company}.")
            except Exception as outlook_error:
                print(f"Outlook fallback error for {company}: {outlook_error}")
                print("Please check your Office365 SMTP settings or use an app password.")
                raise


def main() -> None:
    args = parse_args()
    start_time = datetime.now()
    print(f"Automation started Flexcube_TrialBalRpt_vs_Oracle_GL_Report: {start_time.strftime('%Y-%m-%d %H:%M:%S.%f')}")
    root = Path(args.root)
    base = month_root(root, args.year, args.month)
    default_template = root / "Template_format" / "Template.xlsx"
    # Template ko read karne ki zarurat nahi (purely new workbook generate hota hai)

    template_path: Path = Path(args.template) if args.template else default_template

    output_dir: Path = Path(args.output_dir) if args.output_dir else base / "Reconciliation_Output"

    companies: Iterable[str]
    companies = ["CGCL", "CGHFL"] if args.company == "ALL" else [args.company]

    outputs_by_company: dict[str, Path] = {}
    all_dates_by_company: dict[str, list[str]] = {}
    diff_dates_by_company: dict[str, list[str]] = {}
    for company in companies:
        try:
            output_path = run_company(root, args.year, args.month, output_dir, company)
            outputs_by_company[company] = output_path
            all_dates, diff_dates = find_dates_with_diff(output_path)
            all_dates_by_company[company] = all_dates
            diff_dates_by_company[company] = diff_dates
        except FileNotFoundError as exc:
            print(f"Skipped {company}: {exc}")
            all_dates_by_company[company] = []
            diff_dates_by_company[company] = []

    print("Reconciliation files created:")
    for output in outputs_by_company.values():
        print(output)

    if outputs_by_company and not args.no_email:
        report_date = datetime.now().strftime("%d-%m-%Y")
        send_reconciliation_email(outputs_by_company, report_date, all_dates_by_company, diff_dates_by_company)


if __name__ == "__main__":
    main()
